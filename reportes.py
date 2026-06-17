# ==============================================================================
# INSTRUCCIONES DE INSTALACIÓN Y EMPAQUETADO
# ==============================================================================
# pip install openpyxl pandas
# pyinstaller --noconsole --onefile --name="Galeria_Inventario" app.py
# ==============================================================================

# ============================================
# SECCIÓN 1: IMPORTACIONES Y CONFIGURACIÓN
# ============================================
import os
import datetime
from datetime import timedelta
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl

DB_MAESTRO_DEFAULT = "inventario_maestro.xlsx"
DB_REPORTES_DEFAULT = f"reportes_ventas_{datetime.datetime.now().year}.xlsx"

# ============================================
# SECCIÓN 2: ESTILOS Y DISEÑO DE INTERFAZ
# ============================================
COLOR_PRIMARY = "#2C3E50"    
COLOR_SECONDARY = "#16A085"  
COLOR_BG = "#F8F9FA"         
COLOR_TEXT = "#34495E"       
COLOR_ACCENT = "#E74C3C"     

FONT_TITLE = ("Segoe UI", 14, "bold")
FONT_SUBTITLE = ("Segoe UI", 11, "bold")
FONT_BODY = ("Segoe UI", 10)

def aplicar_estilos_modernos():
    style = ttk.Style()
    style.theme_use('clam')
    style.configure(".", background=COLOR_BG, foreground=COLOR_TEXT, font=FONT_BODY)
    style.configure("TLabel", background=COLOR_BG, foreground=COLOR_TEXT)
    style.configure("TFrame", background=COLOR_BG)
    style.configure("TNotebook", background=COLOR_BG, borderwidth=0)
    style.configure("TNotebook.Tab", background="#BDC3C7", foreground=COLOR_TEXT, padding=[15, 5], font=FONT_BODY)
    style.map("TNotebook.Tab", background=[("selected", COLOR_PRIMARY)], foreground=[("selected", "white")])
    style.configure("TButton", background=COLOR_PRIMARY, foreground="white", borderwidth=0, padding=6, focuscolor=COLOR_PRIMARY)
    style.map("TButton", background=[("active", COLOR_SECONDARY)])
    style.configure("Accent.TButton", background=COLOR_SECONDARY, foreground="white")
    style.map("Accent.TButton", background=[("active", COLOR_PRIMARY)])
    style.configure("TEntry", fieldbackground="white", borderwidth=1)

# ============================================
# SECCIÓN 3: LÓGICA DE NEGOCIO (ADAPTADA A TU EXCEL)
# ============================================

def inicializar_excel_maestro(ruta):
    """Crea la estructura exacta basada en tu captura si el archivo no existe."""
    if not os.path.exists(ruta):
        df = pd.DataFrame(columns=[
            "Codigo de producto", "Departamento", "Origen", "Nombre", 
            "Descripcion", "Precio por unidad", "Cantidad en existencias", 
            "Valor de inventario", "¿Suspendido?"
        ])
        df.to_excel(ruta, index=False)

def buscar_producto_maestro(ruta, codigo):
    """Busca por 'Codigo de producto'."""
    if not os.path.exists(ruta):
        return None
    try:
        df = pd.read_excel(ruta, dtype={"Codigo de producto": str})
        res = df[df["Codigo de producto"] == str(codigo).strip()]
        if not res.empty:
            return res.iloc[0].to_dict()
    except Exception as e:
        messagebox.showerror("Error de lectura", f"No se pudo leer el maestro:\n{e}")
    return None

def guardar_o_actualizar_producto(ruta, item_dict):
    """Inserta nueva fila respetando la estructura completa de tus columnas."""
    inicializar_excel_maestro(ruta)
    try:
        df = pd.read_excel(ruta, dtype={"Codigo de producto": str})
        item_dict["Codigo de producto"] = str(item_dict["Codigo de producto"]).strip()
        
        # Eliminar duplicado si ya existía el código
        df = df[df["Codigo de producto"] != item_dict["Codigo de producto"]]
        
        new_row = pd.DataFrame([item_dict])
        df = pd.concat([df, new_row], ignore_index=True)
        
        # Mantener orden por Departamento
        df.sort_values(by=["Departamento", "Codigo de producto"], inplace=True)
        df.to_excel(ruta, index=False)
        return True
    except Exception as e:
        messagebox.showerror("Error de escritura", f"No se pudo guardar en maestro:\n{e}")
        return False

def modificar_stock_maestro(ruta, codigo, cantidad_cambio):
    """Modifica el stock recalculando automáticamente la columna 'Valor de inventario'."""
    try:
        df = pd.read_excel(ruta, dtype={"Codigo de producto": str})
        codigo_str = str(codigo).strip()
        
        if codigo_str in df["Codigo de producto"].values:
            idx = df[df["Codigo de producto"] == codigo_str].index[0]
            
            # Obtener y calcular stock
            cant_actual = df.at[idx, "Cantidad en existencias"]
            # Validar si viene vacío/NaN en el Excel
            if pd.isna(cant_actual): cant_actual = 0
            
            nueva_cant = int(cant_actual) + cantidad_cambio
            if nueva_cant < 0:
                return False, "La cantidad en existencias no puede ser negativa."
                
            df.at[idx, "Cantidad en existencias"] = nueva_cant
            
            # Recalcular dinámicamente el valor de inventario (Precio * Cantidad)
            precio = df.at[idx, "Precio por unidad"]
            if pd.isna(precio): precio = 0.0
            df.at[idx, "Valor de inventario"] = float(precio) * nueva_cant
            
            df.to_excel(ruta, index=False)
            return True, nueva_cant
        return False, "Código de producto no encontrado."
    except Exception as e:
        return False, str(e)

def obtener_semana_y_mes_asignado(fecha):
    lunes = fecha - timedelta(days=fecha.weekday())
    domingo = lunes + timedelta(days=6)
    if lunes.month == domingo.month:
        mes_asignado = lunes.strftime("%B").capitalize()
    else:
        dias_mes_lunes = (datetime.date(lunes.year, lunes.month, 1) + timedelta(days=32)).replace(day=1) - lunes
        mes_asignado = lunes.strftime("%B").capitalize() if dias_mes_lunes.days >= 4 else domingo.strftime("%B").capitalize()
            
    meses_es = {
        "January": "Enero", "February": "Febrero", "March": "Marzo", "April": "Abril",
        "May": "Mayo", "June": "Junio", "July": "Julio", "August": "Agosto",
        "September": "Septiembre", "October": "Octubre", "November": "Noviembre", "December": "Diciembre"
    }
    return meses_es.get(mes_asignado, mes_asignado), lunes, domingo

def registrar_venta_en_reporte(ruta_reporte, venta_data):
    fecha_v = datetime.datetime.strptime(venta_data["Fecha"], "%Y-%m-%d").date()
    mes_hoja, lunes, domingo = obtener_semana_y_mes_asignado(fecha_v)
    
    if not os.path.exists(ruta_reporte):
        wb = openpyxl.Workbook()
        wb.save(ruta_reporte)
        
    try:
        wb = openpyxl.load_workbook(ruta_reporte)
        if mes_hoja not in wb.sheetnames:
            ws = wb.create_sheet(title=mes_hoja)
            ws.append(["Fecha", "Semana (Rango)", "Código/Nombre", "Tipo", "Cantidad", "Efectivo", "Tarjeta", "Total Descuento/Modif", "Total Dinero"])
        else:
            ws = wb[mes_hoja]
            
        semana_str = f"{lunes.strftime('%d/%m')} al {domingo.strftime('%d/%m')}"
        ws.append([
            venta_data["Fecha"], semana_str, venta_data["Identificador"],
            venta_data["Tipo"], venta_data["Cantidad"], venta_data["Efectivo"],
            venta_data["Tarjeta"], venta_data["Descuento_Aplicado"], venta_data["Total"]
        ])
        
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            wb.remove(wb["Sheet"])
            
        wb.save(ruta_reporte)
        return True
    except Exception as e:
        messagebox.showerror("Error al reportar venta", f"Cierre el archivo de reportes antes de continuar.\n{e}")
        return False

# ============================================
# SECCIÓN 4: CONTROLADORES DE EVENTOS Y UI
# ============================================
class AppGaleria(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sistema de Gestión de Galería de Arte (Columnas Reales)")
        self.geometry("950://680")
        self.geometry("950x680")
        self.configure(bg=COLOR_BG)
        
        self.ruta_maestro = tk.StringVar(value=DB_MAESTRO_DEFAULT)
        self.ruta_reportes = tk.StringVar(value=DB_REPORTES_DEFAULT)
        
        aplicar_estilos_modernos()
        self.crear_interfaz_seleccion_archivo()
        self.crear_cuerpo_principal()
        
        inicializar_excel_maestro(self.ruta_maestro.get())

    def crear_interfaz_seleccion_archivo(self):
        frame_top = ttk.LabelFrame(self, text=" Configuración de Archivos Base (BBDD Excel) ", padding=10)
        frame_top.pack(fill="x", padx=15, pady=10)
        
        ttk.Label(frame_top, text="Excel Maestro (Inventario):").grid(row=0, column=0, sticky="w", pady=2)
        ttk.Entry(frame_top, textvariable=self.ruta_maestro, width=60).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(frame_top, text="Examinar", command=self.examinar_maestro).grid(row=0, column=2, padx=2, pady=2)
        
        ttk.Label(frame_top, text="Excel Reportes Mensuales:").grid(row=1, column=0, sticky="w", pady=2)
        ttk.Entry(frame_top, textvariable=self.ruta_reportes, width=60).grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(frame_top, text="Examinar", command=self.examinar_reportes).grid(row=1, column=2, padx=2, pady=2)

    def examinar_maestro(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
        if ruta:
            self.ruta_maestro.set(ruta)
            inicializar_excel_maestro(ruta)

    def examinar_reportes(self):
        ruta = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
        if ruta: self.ruta_reportes.set(ruta)

    def crear_cuerpo_principal(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.tab_ventas = ttk.Frame(self.notebook, padding=10)
        self.tab_inventario = ttk.Frame(self.notebook, padding=10)
        self.tab_reportes = ttk.Frame(self.notebook, padding=10)
        
        self.notebook.add(self.tab_ventas, text="Salida por Ventas")
        self.notebook.add(self.tab_inventario, text="Ajustes de Inventario")
        self.notebook.add(self.tab_reportes, text="Reportes y Ganancias")
        
        self.disenar_tab_ventas()
        self.disenar_tab_inventario()
        self.disenar_tab_reportes()

    def disenar_tab_ventas(self):
        frame_tipo = ttk.LabelFrame(self.tab_ventas, text=" Tipo de Producto ", padding=10)
        frame_tipo.pack(fill="x", pady=5)
        
        self.var_tipo_pieza = tk.StringVar(value="Registrada")
        ttk.Radiobutton(frame_tipo, text="Pieza Registrada (Con Código)", variable=self.var_tipo_pieza, value="Registrada", command=self.alternar_modo_venta).grid(row=0, column=0, padx=10)
        ttk.Radiobutton(frame_tipo, text="Pieza NO Registrada (Venta Rápida)", variable=self.var_tipo_pieza, value="NoRegistrada", command=self.alternar_modo_venta).grid(row=0, column=1, padx=10)
        
        self.frame_form_ventas = ttk.LabelFrame(self.tab_ventas, text=" Datos de la Transacción ", padding=15)
        self.frame_form_ventas.pack(fill="both", expand=True, pady=10)
        
        self.lbl_identificador = ttk.Label(self.frame_form_ventas, text="Código de Producto:")
        self.lbl_identificador.grid(row=0, column=0, sticky="w", pady=5)
        self.ent_identificador = ttk.Entry(self.frame_form_ventas, width=25)
        self.ent_identificador.grid(row=0, column=1, sticky="w", padx=5, pady=5)
        
        self.btn_buscar_p = ttk.Button(self.frame_form_ventas, text="Buscar Código", command=self.evento_buscar_codigo_venta)
        self.btn_buscar_p.grid(row=0, column=2, sticky="w", padx=5)
        
        self.lbl_info_prod = ttk.Label(self.frame_form_ventas, text="", font=("Segoe UI", 9, "italic"), foreground=COLOR_SECONDARY)
        self.lbl_info_prod.grid(row=0, column=3, columnspan=2, sticky="w")

        ttk.Label(self.frame_form_ventas, text="Cant. Vendida en Efectivo:").grid(row=1, column=0, sticky="w", pady=5)
        self.ent_cant_efectivo = ttk.Entry(self.frame_form_ventas, width=10)
        self.ent_cant_efectivo.insert(0, "0")
        self.ent_cant_efectivo.grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        ttk.Label(self.frame_form_ventas, text="Cant. Vendida con Tarjeta:").grid(row=1, column=2, sticky="w", pady=5)
        self.ent_cant_tarjeta = ttk.Entry(self.frame_form_ventas, width=10)
        self.ent_cant_tarjeta.insert(0, "0")
        self.ent_cant_tarjeta.grid(row=1, column=3, sticky="w", padx=5, pady=5)

        self.lbl_precio_manual = ttk.Label(self.frame_form_ventas, text="Precio Unitario ($):")
        self.lbl_precio_manual.grid(row=2, column=0, sticky="w", pady=5)
        self.ent_precio_manual = ttk.Entry(self.frame_form_ventas, width=15)
        self.ent_precio_manual.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        self.var_tiene_desc = tk.BooleanVar(value=False)
        self.chk_desc = ttk.Checkbutton(self.frame_form_ventas, text="¿Modificar precio o aplicar descuento?", variable=self.var_tiene_desc, command=self.alternar_campos_descuento)
        self.chk_desc.grid(row=3, column=0, columnspan=2, sticky="w", pady=8)
        
        self.lbl_desc = ttk.Label(self.frame_form_ventas, text="Descuento (%) O Nuevo Precio ($):")
        self.lbl_desc.grid(row=4, column=0, sticky="w", pady=5)
        self.ent_desc = ttk.Entry(self.frame_form_ventas, width=15)
        self.ent_desc.grid(row=4, column=1, sticky="w", padx=5, pady=5)
        
        self.alternar_modo_venta()
        
        btn_procesar = ttk.Button(self.tab_ventas, text="PROCESAR Y REGISTRAR VENTA", style="Accent.TButton", command=self.evento_procesar_venta)
        btn_procesar.pack(fill="x", pady=5)

    def alternar_modo_venta(self):
        modo = self.var_tipo_pieza.get()
        if modo == "Registrada":
            self.lbl_identificador.config(text="Código de Producto:")
            self.btn_buscar_p.grid()
            self.lbl_precio_manual.grid_remove()
            self.ent_precio_manual.grid_remove()
            self.chk_desc.grid()
            self.alternar_campos_descuento()
        else:
            self.lbl_identificador.config(text="Nombre del Producto:")
            self.btn_buscar_p.grid_remove()
            self.lbl_info_prod.config(text="")
            self.lbl_precio_manual.grid()
            self.ent_precio_manual.grid()
            self.chk_desc.grid_remove()
            self.lbl_desc.grid_remove()
            self.ent_desc.grid_remove()

    def alternar_campos_descuento(self):
        if self.var_tiene_desc.get() and self.var_tipo_pieza.get() == "Registrada":
            self.lbl_desc.grid()
            self.ent_desc.grid()
        else:
            self.lbl_desc.grid_remove()
            self.ent_desc.grid_remove()

    def evento_buscar_codigo_venta(self):
        cod = self.ent_identificador.get().strip()
        if not cod: return
        prod = buscar_producto_maestro(self.ruta_maestro.get(), cod)
        if prod:
            self.lbl_info_prod.config(text=f"Encontrado: {prod['Nombre']} | Existencia: {prod['Cantidad en existencias']} | Precio: ${prod['Precio por unidad']}", foreground=COLOR_SECONDARY)
        else:
            self.lbl_info_prod.config(text="Producto no localizado en el Maestro.", foreground=COLOR_ACCENT)

    def evento_procesar_venta(self):
        modo = self.var_tipo_pieza.get()
        identificador = self.ent_identificador.get().strip()
        
        try:
            cant_ef = int(self.ent_cant_efectivo.get())
            cant_tj = int(self.ent_cant_tarjeta.get())
        except ValueError:
            messagebox.showerror("Error", "Las cantidades deben ser enteros.")
            return
            
        cant_total = cant_ef + cant_tj
        if cant_total <= 0:
            messagebox.showerror("Error", "La cantidad total debe ser mayor a 0.")
            return

        precio_unitario = 0.0
        descuento_nota = "Ninguno"
        
        if modo == "Registrada":
            prod = buscar_producto_maestro(self.ruta_maestro.get(), identificador)
            if not prod:
                messagebox.showerror("Error", "Código no válido.")
                return
                
            if prod["Cantidad en existencias"] < cant_total:
                messagebox.showerror("Error", f"Stock insuficiente. Disponible: {prod['Cantidad en existencias']}")
                return
                
            precio_unitario = float(prod["Precio por unidad"])
            
            if self.var_tiene_desc.get():
                val_desc = self.ent_desc.get().strip()
                try:
                    if "%" in val_desc:
                        porcentaje = float(val_desc.replace("%", ""))
                        precio_unitario = precio_unitario * (1 - (porcentaje / 100))
                        descuento_nota = f"Desc {porcentaje}%"
                    else:
                        precio_unitario = float(val_desc)
                        descuento_nota = f"Precio Modif a ${precio_unitario}"
                except ValueError:
                    messagebox.showerror("Error", "Formato de descuento incorrecto (Ej: 10% o 5000).")
                    return
                    
            modificar_stock_maestro(self.ruta_maestro.get(), identificador, -cant_total)
            
        else:
            if not identificador:
                messagebox.showerror("Error", "Escriba el nombre del producto.")
                return
            try:
                precio_unitario = float(self.ent_precio_manual.get())
            except ValueError:
                messagebox.showerror("Error", "Precio unitario inválido.")
                return
            descuento_nota = "Venta Rápida"

        tot_efectivo = cant_ef * precio_unitario
        tot_tarjeta = cant_tj * precio_unitario
        total_dinero = tot_efectivo + tot_tarjeta
        
        venta_obj = {
            "Fecha": datetime.date.today().strftime("%Y-%m-%d"),
            "Identificador": identificador,
            "Tipo": modo,
            "Cantidad": cant_total,
            "Efectivo": tot_efectivo,
            "Tarjeta": tot_tarjeta,
            "Descuento_Aplicado": descuento_nota,
            "Total": total_dinero
        }
        
        if registrar_venta_en_reporte(self.ruta_reportes.get(), venta_obj):
            messagebox.showinfo("Éxito", f"Venta registrada. Total: ${total_dinero:.2f}")
            self.ent_identificador.delete(0, tk.END)
            self.ent_cant_efectivo.delete(0, tk.END)
            self.ent_cant_efectivo.insert(0, "0")
            self.ent_cant_tarjeta.delete(0, tk.END)
            self.ent_cant_tarjeta.insert(0, "0")
            self.ent_precio_manual.delete(0, tk.END)
            self.ent_desc.delete(0, tk.END)
            self.lbl_info_prod.config(text="")

    def disenar_tab_inventario(self):
        # Ajuste de stock de existentes
        frame_existente = ttk.LabelFrame(self.tab_inventario, text=" Ajustar Stock de Pieza Existente ", padding=10)
        frame_existente.pack(fill="x", pady=5)
        
        ttk.Label(frame_existente, text="Código de Producto:").grid(row=0, column=0, sticky="w", pady=5)
        self.ent_ajuste_cod = ttk.Entry(frame_existente, width=20)
        self.ent_ajuste_cod.grid(row=0, column=1, padx=5)
        
        ttk.Label(frame_existente, text="Cantidad (+/-):").grid(row=0, column=2, sticky="w", pady=5)
        self.ent_ajuste_cant = ttk.Entry(frame_existente, width=10)
        self.ent_ajuste_cant.grid(row=0, column=3, padx=5)
        
        ttk.Button(frame_existente, text="Aplicar Ajuste", command=self.evento_ajustar_stock_existente).grid(row=0, column=4, padx=5)
        
        # Alta de piezas nuevas con todas tus columnas reales
        frame_nueva = ttk.LabelFrame(self.tab_inventario, text=" Registrar Alta de Nueva Pieza (Estructura Real) ", padding=10)
        frame_nueva.pack(fill="both", expand=True, pady=5)
        
        fields = [
            ("Código de Producto:", "Codigo de producto"),
            ("Departamento:", "Departamento"),
            ("Origen (Local/Importado):", "Origen"),
            ("Nombre del Producto:", "Nombre"),
            ("Descripción / Dimensiones:", "Descripcion"),
            ("Precio por Unidad ($):", "Precio por unidad"),
            ("Cantidad Inicial:", "Cantidad en existencias"),
            ("¿Suspendido? (Si/No):", "¿Suspendido?")
        ]
        
        self.dict_alta_entries = {}
        for idx, (label_text, key) in enumerate(fields):
            ttk.Label(frame_nueva, text=label_text).grid(row=idx, column=0, sticky="w", pady=4, padx=5)
            entry = ttk.Entry(frame_nueva, width=40)
            entry.grid(row=idx, column=1, sticky="w", pady=4, padx=5)
            if key == "¿Suspendido?": entry.insert(0, "No")
            self.dict_alta_entries[key] = entry
            
        ttk.Button(frame_nueva, text="GUARDAR NUEVA PIEZA EN EXCEL MAESTRO", style="Accent.TButton", command=self.evento_alta_nueva_pieza).grid(row=len(fields), column=0, columnspan=2, pady=15)

    def evento_ajustar_stock_existente(self):
        cod = self.ent_ajuste_cod.get().strip()
        try:
            cambio = int(self.ent_ajuste_cant.get())
        except ValueError:
            messagebox.showerror("Error", "La cantidad debe ser un entero.")
            return
            
        exito, msg = modificar_stock_maestro(self.ruta_maestro.get(), cod, cambio)
        if exito:
            messagebox.showinfo("Éxito", f"Inventario modificado. Nueva existencia: {msg}")
            self.ent_ajuste_cod.delete(0, tk.END)
            self.ent_ajuste_cant.delete(0, tk.END)
        else:
            messagebox.showerror("Error", msg)

    def evento_alta_nueva_pieza(self):
        codigo = self.dict_alta_entries["Codigo de producto"].get().strip()
        depto = self.dict_alta_entries["Departamento"].get().strip()
        origen = self.dict_alta_entries["Origen"].get().strip()
        nombre = self.dict_alta_entries["Nombre"].get().strip()
        desc = self.dict_alta_entries["Descripcion"].get().strip()
        suspendido = self.dict_alta_entries["¿Suspendido?"].get().strip()
        
        if not (codigo and depto and nombre):
            messagebox.showerror("Error", "Código, Departamento y Nombre son obligatorios.")
            return
            
        try:
            precio = float(self.dict_alta_entries["Precio por unidad"].get())
            cantidad = int(self.dict_alta_entries["Cantidad en existencias"].get())
        except ValueError:
            messagebox.showerror("Error", "Precio y Cantidad deben ser numéricos.")
            return
            
        if buscar_producto_maestro(self.ruta_maestro.get(), codigo):
            messagebox.showerror("Error", f"El código '{codigo}' ya existe.")
            return
            
        item_dict = {
            "Codigo de producto": codigo,
            "Departamento": depto,
            "Origen": origen,
            "Nombre": nombre,
            "Descripcion": desc,
            "Precio por unidad": precio,
            "Cantidad en existencias": cantidad,
            "Valor de inventario": precio * cantidad,
            "¿Suspendido?": suspendido
        }
        
        if guardar_o_actualizar_producto(self.ruta_maestro.get(), item_dict):
            messagebox.showinfo("Éxito", "Producto agregado de acuerdo a las columnas de tu Excel.")
            for k, entry in self.dict_alta_entries.items():
                entry.delete(0, tk.END)
                if k == "¿Suspendido?": entry.insert(0, "No")

    def disenar_tab_reportes(self):
        frame_filtro = ttk.LabelFrame(self.tab_reportes, text=" Selección de Semana para Análisis ", padding=10)
        frame_filtro.pack(fill="x", pady=5)
        
        ttk.Label(frame_filtro, text="Fecha (AAAA-MM-DD):").grid(row=0, column=0, sticky="w")
        self.ent_fecha_reporte = ttk.Entry(frame_filtro, width=15)
        self.ent_fecha_reporte.insert(0, datetime.date.today().strftime("%Y-%m-%d"))
        self.ent_fecha_reporte.grid(row=0, column=1, padx=5)
        
        ttk.Button(frame_filtro, text="Generar Reporte Integral", command=self.evento_calcular_reporte_semanal).grid(row=0, column=2, padx=5)
        
        self.lbl_semana_activa = ttk.Label(frame_filtro, text="Semana Evaluada: -", font=("Segoe UI", 10, "bold"), foreground=COLOR_SECONDARY)
        self.lbl_semana_activa.grid(row=1, column=0, columnspan=3, pady=5, sticky="w")

        self.canvas_rep = tk.Canvas(self.tab_reportes, bg=COLOR_BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_reportes, orient="vertical", command=self.canvas_rep.yview)
        self.scrollable_frame = ttk.Frame(self.canvas_rep)
        
        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas_rep.configure(scrollregion=self.canvas_rep.bbox("all")))
        self.canvas_rep.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas_rep.configure(yscrollcommand=scrollbar.set)
        
        self.canvas_rep.pack(side="left", fill="both", expand=True, pady=5)
        scrollbar.pack(side="right", fill="y")
        
        self.frame_t1 = ttk.LabelFrame(self.scrollable_frame, text=" Tabla 1: Ventas de Piezas Registradas ", padding=5)
        self.frame_t1.pack(fill="x", expand=True, pady=5)
        self.frame_t2 = ttk.LabelFrame(self.scrollable_frame, text=" Tabla 2: Ventas de Piezas NO Registradas ", padding=5)
        self.frame_t2.pack(fill="x", expand=True, pady=5)
        self.frame_totales = ttk.LabelFrame(self.scrollable_frame, text=" Resumen Financiero y Comisión de Ventas ", padding=10)
        self.frame_totales.pack(fill="x", expand=True, pady=5)

    def evento_calcular_reporte_semanal(self):
        try:
            fecha_ref = datetime.datetime.strptime(self.ent_fecha_reporte.get().strip(), "%Y-%m-%d").date()
        except ValueError:
            messagebox.showerror("Error", "Formato de fecha inválido.")
            return
            
        mes_hoja, lunes, domingo = obtener_semana_y_mes_asignado(fecha_ref)
        semana_str = f"{lunes.strftime('%d/%m')} al {domingo.strftime('%d/%m')}"
        self.lbl_semana_activa.config(text=f"Semana Evaluada: Del {lunes.strftime('%d-%m-%Y')} al {domingo.strftime('%d-%m-%Y')} ({mes_hoja})")
        
        ruta_rep = self.ruta_reportes.get()
        if not os.path.exists(ruta_rep):
            messagebox.showwarning("Sin Datos", "No existe archivo de reportes.")
            return
            
        try:
            df = pd.read_excel(ruta_rep, sheet_name=mes_hoja)
        except Exception:
            df = pd.DataFrame()
            
        if df.empty:
            messagebox.showinfo("Vacío", "Sin registros este mes.")
            return
            
        df_semana = df[df["Semana (Rango)"] == semana_str]
        if df_semana.empty:
            messagebox.showinfo("Vacío", "Sin transacciones esta semana.")
            return
            
        for w in self.frame_t1.winfo_children(): w.destroy()
        for w in self.frame_t2.winfo_children(): w.destroy()
        for w in self.frame_totales.winfo_children(): w.destroy()
        
        df_reg = df_semana[df_semana["Tipo"] == "Registrada"]
        ttk.Label(self.frame_t1, text="Código | Cantidad Vendida | Total ($)", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        if not df_reg.empty:
            t1_g = df_reg.groupby("Código/Nombre").agg({"Cantidad": "sum", "Total Dinero": "sum"}).reset_index()
            for _, r in t1_g.iterrows():
                ttk.Label(self.frame_t1, text=f"• {r['Código/Nombre']}  ->  {r['Cantidad']} u.  ->  ${r['Total Dinero']:.2f}").pack(anchor="w", padx=10)
        
        df_noreg = df_semana[df_semana["Tipo"] == "NoRegistrada"]
        ttk.Label(self.frame_t2, text="Nombre Producto | Cantidad Vendida | Total ($)", font=("Segoe UI", 9, "bold")).pack(anchor="w")
        if not df_noreg.empty:
            t2_g = df_noreg.groupby("Código/Nombre").agg({"Cantidad": "sum", "Total Dinero": "sum"}).reset_index()
            for _, r in t2_g.iterrows():
                ttk.Label(self.frame_t2, text=f"• {r['Código/Nombre']}  ->  {r['Cantidad']} u.  ->  ${r['Total Dinero']:.2f}").pack(anchor="w", padx=10)
                
        gran_total = df_semana["Total Dinero"].sum()
        top10_df = df_semana.groupby("Código/Nombre").agg({"Cantidad": "sum"}).sort_values(by="Cantidad", ascending=False).head(10)
        
        ttk.Label(self.frame_totales, text=f"TOTAL GENERAL: ${gran_total:.2f}", font=("Segoe UI", 12, "bold"), foreground=COLOR_SECONDARY).grid(row=0, column=0, columnspan=2, sticky="w", pady=5)
        ttk.Label(self.frame_totales, text=f"- Dueña (50%): ${gran_total*0.5:.2f}").grid(row=1, column=0, sticky="w", padx=15)
        ttk.Label(self.frame_totales, text=f"- Guías (25%): ${gran_total*0.25:.2f}").grid(row=2, column=0, sticky="w", padx=15)
        ttk.Label(self.frame_totales, text=f"- Encargada (25%): ${gran_total*0.25:.2f}").grid(row=3, column=0, sticky="w", padx=15)
        
        ttk.Label(self.frame_totales, text="Top 10 Productos:", font=("Segoe UI", 10, "bold")).grid(row=1, column=1, sticky="w", padx=30)
        r_idx = 2
        for name, row in top10_df.iterrows():
            ttk.Label(self.frame_totales, text=f"{r_idx-1}. {name} ({row['Cantidad']} u)").grid(row=r_idx, column=1, sticky="w", padx=45)
            r_idx += 1

if __name__ == "__main__":
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception: pass
    AppGaleria().mainloop()